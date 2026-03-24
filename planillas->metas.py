import streamlit as st
import pandas as pd
import re
import io
import datetime
from openpyxl import load_workbook

st.set_page_config(page_title="Procesador MAGA", page_icon="🌽", layout="wide")

st.title("🌽 Procesador de Beneficiarios MAGA")
st.markdown(
    "Sube uno o varios archivos fuente y los tres archivos de totales. "
    "El sistema calculará los totales desagregados y los escribirá en las hojas correctas."
)

REFERENCE_YEAR = datetime.datetime.now().year

# ─── Intervention code → keyword to find the correct sheet in the output file ───
AGRI_CODE_TO_KW = {
    101: "GRANOS BASICOS",
    102: "HORTALIZAS",
    103: "FRUTALES",
    104: "CULTIVOS PERMANENTES",
    105: "SAF",
    107: "ASOCIATIVIDAD",
    108: "ESTRUCTURAS DE CONSERVACION",
    109: "PRACTICA",
    110: "ABONO ORGANICO",
    111: "FERTILIZANTE",
    112: "BIOCIDAS",
}
AGRI_BANCO_KW = "BANCO COMUNITARIO"   # no numeric code; matched by description text

PECUARIO_CODE_TO_KW = {
    201: "BOVINOS",
    202: "PORCINOS",
    203: "CAPRINOS",
    204: "AVES DE POSTURA",
    205: "PISCICOLA",
    206: "APICOLA",
    106: "AUTOGESTION",
    107: "ASOCIATIVIDAD",
}

HR_CODE_TO_KW = {
    301: "PRODUCCION DE PATIO",
    302: "SEGURIDAD ALIMENTARIA",
    303: "MEJORAMIENTO DEL HOGAR",
    304: "SALUD E HIGIENE",
    305: "EMPRENDIMIENTOS",
}


# ─── Helpers ────────────────────────────────────────────────────────────────

def normalize(s):
    if pd.isna(s):
        return ""
    return re.sub(r"\s+", " ", str(s).upper().strip())


def extract_code(desc):
    """Extract the 3-digit category code from a description like 'HORTALIZAS -102- : ...'"""
    if pd.isna(desc):
        return None
    m = re.search(r"-(\d{3})\d*-", str(desc))
    return int(m.group(1)) if m else None


def get_age_group(birth_val):
    if pd.isna(birth_val):
        return None
    try:
        by = birth_val.year if hasattr(birth_val, "year") else int(birth_val)
    except Exception:
        return None
    age = REFERENCE_YEAR - by
    if 13 <= age <= 18:   return "13_18"
    if 19 <= age <= 30:   return "19_30"
    if 31 <= age <= 59:   return "31_59"
    if age >= 60:         return "60p"
    return None


def get_linguistic(val):
    v = normalize(val)
    if "MAM" in v:                                          return "mam"
    if "KICHE" in v or "K'ICHE" in v or "KICH" in v:       return "kiche"
    return "ladino"   # Español, Ladino, Extranjero, unknown → ladino bucket


def get_pueblo(ethnic):
    return "maya" if "MAYA" in normalize(ethnic) else "ladino"


def find_sheet_by_kw(wb, keyword):
    kw = keyword.upper()
    for name in wb.sheetnames:
        if kw in name.upper():
            return wb[name]
    return None


def find_municipality_row(ws, municipality):
    """Find the Excel row number for a given municipality in column C (1-indexed)."""
    muni_norm = normalize(municipality)
    for row in ws.iter_rows(min_row=5, max_col=3, values_only=False):
        cell = row[2]   # column C
        if normalize(cell.value) == muni_norm:
            return cell.row
    return None


# ─── Source file reader ─────────────────────────────────────────────────────

def read_source_sheet(file_bytes, sheet_name):
    """
    Read one source sheet.  Header is on row 7 (index 6).
    Returns a cleaned DataFrame or None if the sheet is missing / empty.
    """
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name,
                           header=6, dtype=str)
    except Exception:
        return None

    if df.empty:
        return None

    # The columns we need by positional index (0-based after header row):
    # 5=Sexo, 6=Año nacimiento, 8=Municipio, 10=Étnica, 11=Lingüística, 13=Descripción
    try:
        df = df.rename(columns={
            df.columns[5]:  "sexo",
            df.columns[6]:  "birth_year",
            df.columns[8]:  "municipio",
            df.columns[10]: "etnia",
            df.columns[11]: "linguistica",
            df.columns[13]: "descripcion",
        })
    except IndexError:
        return None

    df["birth_year"] = pd.to_datetime(df["birth_year"], errors="coerce")
    # Drop rows that have neither a municipality nor a sex value (template filler rows)
    df = df[df["municipio"].notna() & df["municipio"].str.strip().ne("")]
    return df if not df.empty else None


# ─── Row → sheet keyword mapper ─────────────────────────────────────────────

def assign_sheet_kw(df, code_map, banco_kw=None):
    """
    Add '_sheet_kw' column.  Returns (tagged_df, list_of_unmatched_descriptions).
    """
    unmatched = []

    def mapper(desc):
        d = str(desc) if not pd.isna(desc) else ""
        # Special case: Bancos comunitarios de semillas (no numeric code)
        if banco_kw and "BANCO" in d.upper() and "SEMILLA" in d.upper():
            return banco_kw
        code = extract_code(d)
        if code is not None and code in code_map:
            return code_map[code]
        unmatched.append(d)
        return None

    df = df.copy()
    df["_sheet_kw"] = df["descripcion"].apply(mapper)
    return df, list(set(unmatched))


# ─── Aggregation ────────────────────────────────────────────────────────────

def aggregate(df):
    """
    Group rows by (sheet_kw, municipality) and count disaggregated beneficiaries.
    Returns { sheet_kw: { municipality_upper: counts_dict } }
    """
    results = {}

    for _, row in df.iterrows():
        kw   = row.get("_sheet_kw")
        muni = normalize(row.get("municipio", ""))
        sexo = normalize(row.get("sexo", ""))

        if not kw or not muni:
            continue
        is_female = "MUJER" in sexo
        is_male   = "HOMBRE" in sexo
        if not is_female and not is_male:
            continue

        ling   = get_linguistic(row.get("linguistica", ""))
        pueblo = get_pueblo(row.get("etnia", ""))
        age    = get_age_group(row.get("birth_year"))

        if kw not in results:
            results[kw] = {}
        if muni not in results[kw]:
            results[kw][muni] = {
                "f_ladino": 0, "f_mam": 0, "f_kiche": 0,
                "f_13_18": 0, "f_19_30": 0, "f_31_59": 0, "f_60p": 0,
                "m_ladino": 0, "m_mam": 0, "m_kiche": 0,
                "m_13_18": 0, "m_19_30": 0, "m_31_59": 0, "m_60p": 0,
                "maya": 0, "ladino_total": 0,
            }
        c = results[kw][muni]
        prefix = "f" if is_female else "m"

        c[f"{prefix}_{ling}"] += 1
        if age:
            c[f"{prefix}_{age}"] += 1
        if pueblo == "maya":
            c["maya"] += 1
        else:
            c["ladino_total"] += 1

    return results


# ─── Column layout detection ─────────────────────────────────────────────────

def detect_layout(ws):
    """
    Read header rows 1-3 of a totals sheet and return a dict mapping
    field names to column indices (1-based).  Returns None on failure.
    """
    # Collect row-3 labels
    row3 = {cell.column: normalize(cell.value) for cell in ws[3] if cell.value}

    # Find where MUJERES / HOMBRES blocks start (rows 1 or 2)
    female_start = male_start = None
    for r in [1, 2]:
        for cell in ws[r]:
            v = normalize(cell.value)
            if "MUJERES" in v and female_start is None:
                female_start = cell.column
            if "HOMBRES" in v and male_start is None:
                male_start = cell.column

    if female_start is None or male_start is None:
        return None

    layout = {}
    age_map = {
        "13 - 18": "13_18", "13  -  18": "13_18",
        "19  -  30": "19_30", "19 - 30": "19_30",
        "31  -  59": "31_59", "31 - 59": "31_59",
        "MAYOR A 60": "60p",
    }

    for col, val in sorted(row3.items()):
        in_f = female_start <= col < male_start
        in_m = col >= male_start

        prefix = ("f" if in_f else "m") if (in_f or in_m) else None
        if prefix is None:
            continue

        if "LADINO" in val:
            key = f"{prefix}_ladino"
            # After MAYA in the male block it becomes the pueblo-level ladino total
            if in_m and "maya" in layout and f"m_ladino" in layout:
                key = "ladino_total"
            if key not in layout:
                layout[key] = col

        elif "MAM" in val and f"{prefix}_mam" not in layout:
            layout[f"{prefix}_mam"] = col

        elif ("KICHE" in val or "K'ICHE" in val) and f"{prefix}_kiche" not in layout:
            layout[f"{prefix}_kiche"] = col

        elif "MAYA" in val and "maya" not in layout:
            layout["maya"] = col

        else:
            for lbl, age_key in age_map.items():
                if lbl in val:
                    full_key = f"{prefix}_{age_key}"
                    if full_key not in layout:
                        layout[full_key] = col
                    break

    return layout if len(layout) >= 6 else None


# ─── Write to workbook ───────────────────────────────────────────────────────

def write_to_workbook(totals_bytes, aggregated, warnings):
    """
    Load a totals workbook, increment cells for each (sheet_kw, municipality),
    return the modified workbook as bytes.
    """
    wb = load_workbook(io.BytesIO(totals_bytes))

    for sheet_kw, muni_data in aggregated.items():
        ws = find_sheet_by_kw(wb, sheet_kw)
        if ws is None:
            warnings.append(f"⚠️ Hoja con keyword **'{sheet_kw}'** no encontrada en el archivo de totales.")
            continue

        layout = detect_layout(ws)
        if layout is None:
            warnings.append(f"⚠️ No se pudo leer el encabezado de la hoja **'{ws.title}'**. Se omitió.")
            continue

        for muni, counts in muni_data.items():
            row_num = find_municipality_row(ws, muni)
            if row_num is None:
                warnings.append(f"⚠️ Municipio **'{muni}'** no encontrado en hoja **'{ws.title}'**. Se omitió.")
                continue

            for field, value in counts.items():
                if value == 0:
                    continue
                col = layout.get(field)
                if col is None:
                    continue
                cell = ws.cell(row=row_num, column=col)
                try:
                    existing = int(cell.value) if cell.value else 0
                except (ValueError, TypeError):
                    existing = 0
                cell.value = existing + value

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# ─────────────────────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────────────────────

st.header("1. Archivos fuente")
source_files = st.file_uploader(
    "Sube uno o más archivos fuente (un archivo por municipio)",
    type=["xlsx"],
    accept_multiple_files=True,
)

st.header("2. Archivos de totales")
c1, c2, c3 = st.columns(3)
with c1:
    hr_file   = st.file_uploader("Totales HOGAR RURAL",  type=["xlsx"], key="hr")
with c2:
    agri_file = st.file_uploader("Totales AGRÍCOLA",     type=["xlsx"], key="agri")
with c3:
    pec_file  = st.file_uploader("Totales PECUARIO",     type=["xlsx"], key="pec")

st.divider()

run = st.button("▶️  Procesar", type="primary",
                disabled=(not source_files) or not (hr_file or agri_file or pec_file))

if run:
    warnings    = []
    all_unmatched = []
    agri_agg    = {}
    pec_agg     = {}
    hr_agg      = {}

    progress = st.progress(0, text="Leyendo archivos fuente…")

    for i, src in enumerate(source_files):
        src_bytes = src.read()
        fname = src.name

        for sheet_name, code_map, banco_kw, accumulator in [
            ("AGRICOLA",    AGRI_CODE_TO_KW,     AGRI_BANCO_KW, agri_agg),
            ("PECUARIO",    PECUARIO_CODE_TO_KW, None,          pec_agg),
            ("HOGAR RURAL", HR_CODE_TO_KW,       None,          hr_agg),
        ]:
            df = read_source_sheet(src_bytes, sheet_name)
            if df is None:
                continue

            df, unmatched = assign_sheet_kw(df, code_map, banco_kw)
            for u in unmatched:
                if u.strip():
                    all_unmatched.append(f"`{fname}` / {sheet_name}: `{u}`")

            df_mapped = df[df["_sheet_kw"].notna()]
            if df_mapped.empty:
                continue

            result = aggregate(df_mapped)

            for kw, muni_data in result.items():
                if kw not in accumulator:
                    accumulator[kw] = {}
                for muni, counts in muni_data.items():
                    if muni not in accumulator[kw]:
                        accumulator[kw][muni] = {k: 0 for k in counts}
                    for k, v in counts.items():
                        accumulator[kw][muni][k] += v

        progress.progress((i + 1) / len(source_files), text=f"Procesado: {fname}")

    progress.empty()

    # Write output files
    outputs = {}
    if hr_file and hr_agg:
        outputs["HR_totales_actualizado.xlsx"] = write_to_workbook(
            hr_file.read(), hr_agg, warnings)
    if agri_file and agri_agg:
        outputs["AGRICOLA_totales_actualizado.xlsx"] = write_to_workbook(
            agri_file.read(), agri_agg, warnings)
    if pec_file and pec_agg:
        outputs["PECUARIO_totales_actualizado.xlsx"] = write_to_workbook(
            pec_file.read(), pec_agg, warnings)

    st.success(f"✅ {len(source_files)} archivo(s) fuente procesados.")

    if outputs:
        st.header("3. Descarga los archivos actualizados")
        dl_cols = st.columns(len(outputs))
        for (fname, data), col in zip(outputs.items(), dl_cols):
            col.download_button(
                label=f"⬇️ {fname}",
                data=data,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    else:
        st.warning("No se generaron archivos. Verifica que los archivos fuente contengan datos.")

    if all_unmatched:
        with st.expander(f"⚠️ {len(all_unmatched)} fila(s) ignoradas — código de intervención no reconocido"):
            st.info(
                "Estas filas no se escribieron en ningún archivo de totales porque "
                "su descripción no coincide con ninguna categoría conocida."
            )
            for u in all_unmatched:
                st.markdown(f"- {u}")

    if warnings:
        with st.expander(f"ℹ️ {len(warnings)} aviso(s)"):
            for w in warnings:
                st.markdown(w)
